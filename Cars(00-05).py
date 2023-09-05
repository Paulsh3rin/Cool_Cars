import pandas as pd
import requests
from bs4 import BeautifulSoup
<<<<<<< HEAD
from openpyxl import workbook
from openpyxl.utils.dataframe import dataframe_to_rows 

pd.set_option('display.max_row', 100)
cars_url = "https://www.kbb.com/car-finder/"
parameters = {
    "manufacturers": "bmw|audi|bentley|astonmartin|acura|alfaromeo|volvo|volkswagen|toyota|vinfast|tesla|suzuki|srt|subaru|smart|scion|saturn|saab|rollsroyce|rivian|ram|porsche|polestar|pontiac|plymouth|panoz|nissan|mini|mitsubishi|oldsmobile|buick|cadillac|chevrolet|chrysler|daihatsu|eagle|fiat|ford|genesis|gmc|hummer|infiniti|jaguar|kia|landrover|lincoln|lucid|maybach|mclaren|mercury|mercedesbenz|mazda|maserati|lotus|lexus|lamborghini|jeep|isuzu|hyundai|honda|geo|freightliner|fisker|dodge|ferrari|daewoo",
    "years": "2000-2005",
    "seolink": "false"
}


car_names = []

#iterate through pages
for page_num in range(1, 57):
    parameters["css-1c29x0g e1hywep21"] = str(page_num)
    response = requests.get(cars_url, params=parameters)
    soup = BeautifulSoup(response.content, "html.parser")

#Extract car names from current page and append to list 
    car_names.extend([car.text.strip() for car in soup.find_all('h2')])
    car_types = [a.text.strip() for a in soup.select('.css-3oc9y8.e19qstch15 a')]
=======
 
cars_url = "https://www.kbb.com/car-finder/page-{}/"

car_names = []

session = requests.Session()

#iterate through pages
for page_num in range(1, 57):
    page_url = cars_url.format(page_num)  # Generate the URL for the current page
    
    # Additional parameters
    parameters = {
        "manufacturers": "bmw|audi|bentley|astonmartin|acura|alfaromeo|volvo|volkswagen|toyota|vinfast|tesla|suzuki|srt|subaru|smart|scion|saturn|saab|rollsroyce|rivian|ram|porsche|polestar|pontiac|plymouth|panoz|nissan|mini|mitsubishi|oldsmobile|buick|cadillac|chevrolet|chrysler|daihatsu|eagle|fiat|ford|genesis|gmc|hummer|infiniti|jaguar|kia|landrover|lincoln|lucid|maybach|mclaren|mercury|mercedesbenz|mazda|maserati|lotus|lexus|lamborghini|jeep|isuzu|hyundai|honda|geo|freightliner|fisker|dodge|ferrari|daewoo",
        "years": "2000-2005",
        "seolink": "false"
    }

    response = requests.get(page_url, params=parameters, allow_redirects=False)

    # Check if it's a redirect
    if response.status_code == 302:
        redirect_url = response.headers['Location']
        response = session.get(redirect_url)

    soup = BeautifulSoup(response.content, "html.parser")

    #Extract car names from current page and append to list 
    car_names.extend([car.text.strip() for car in soup.find_all('h2')])
>>>>>>> f6d36b346f53b4b962a7005ffec1f1e4cd407f8d

#Create the dataframe
df = pd.DataFrame({'Model':car_names})

<<<<<<< HEAD
wb = Workbook()
ws = wb.create_sheet('CarNames(2000-2005)')

for r in dataframe_to_rows(df, header=True, index=False):
    ws.append(r)
wb.remove(wb['Sheet'])

wb.save('/workspaces/1202/Project/Code/PlayerData.xlsx')
=======
df.to_excel("CarsData(2000-2005).xlsx", index=False)



>>>>>>> f6d36b346f53b4b962a7005ffec1f1e4cd407f8d


